from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd
import lxml
from flask import send_file
def as_text(value):
    if value is None:
        return ""
    return str(value)

# def export_testing_Report(data, app_name, blank_file, new_file):
#     workbook = load_workbook( filename=blank_file)
#     sheet = workbook.active
#     sheet["D6"] = "TEST"
#     # for row in data:
#     #     sheet["D" + str(start_row)] = app_name
#     #     sheet["E" + str(start_row)] = row[2]
#     #     sheet["F" + str(start_row)] = row[3]
#     #     sheet["G" + str(start_row)] = row[4]
#     #     sheet["H" + str(start_row)] = row[5]
#     #     sheet["I" + str(start_row)] = row[5]
#     #     sheet["J" + str(start_row)] = row[6]
#     #     sheet["K" + str(start_row)] = row[7]
#     #     sheet["L" + str(start_row)] = row[8]
#     #     start_row = start_row + 1
#     # for column_cells in sheet.columns:
#     #     new_column_length = max(len(as_text(cell.value)) for cell in column_cells)
#     #     new_column_letter = (get_column_letter(column_cells[0].column))
#     #     if new_column_length > 0:
#     #         sheet.column_dimensions[new_column_letter].width = new_column_length + 3
#     workbook.save(filename=new_file)

def export_testing_Report(data, app_name, blank_file, new_file):
  test_results = pd.DataFrame(data)
  test_results.columns =['Test ID', 'App ID', 'Date', 'Scenario', 'Duration', 'Source Location', 'Target Location', 'Evidence', 'Result']
  test_results.to_csv(new_file, index=False)

def export_updated_TRAP(data, appname):
  pass

def returnTRAP(trapFile):
  importedDF = pd.read_excel(trapFile, skiprows = 6, nrows=22,  usecols= 'D', header=None)
  trap_json = importedDF.to_json()
  return trap_json

def exportTRAP(data, folder_location):
  to_csv_array = []
  for row in data:
    if len(row) == 2:
      to_csv_array.append(row[1])
    elif len(row) == 1:
      to_csv_array.append(row[0])
  # object_to_export = {
  #   "Pre-Event Planning":[to_csv_array[0], to_csv_array[1], to_csv_array[2]],
  #   "Pre-Declaration":[to_csv_array[3], to_csv_array[4], to_csv_array[5]],
  #   "Source Shutdown":[to_csv_array[6], to_csv_array[7], to_csv_array[8], to_csv_array[9]],
  #   "Starting Target":[to_csv_array[10], to_csv_array[11], to_csv_array[12], to_csv_array[13],to_csv_array[14], to_csv_array[15], to_csv_array[16], to_csv_array[17]],
  # "Environment Verification":[to_csv_array[18], to_csv_array[19], to_csv_array[20], to_csv_array[22]]
  # }
  dataframe = pd.DataFrame(to_csv_array)
  return dataframe.to_csv(folder_location, index=False, header=False)

  
  